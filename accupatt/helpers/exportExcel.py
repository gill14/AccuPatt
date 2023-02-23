import pandas as pd
from accupatt.helpers.dBBridge import load_from_db
from accupatt.models.passData import Pass
from accupatt.models.seriesData import SeriesData
from accupatt.models.sprayCard import SprayCard
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows


def safe_report(filesToInclude: list[str], saveFile: str):
    wb = Workbook()
    # Sheet created by default
    ws: Worksheet = wb.active
    header = [
        "Date",
        "Event",
        "Location",
        "Analyst",
        "Liquid",
        "Remote Analyst",
        "Pilot Last",
        "Pilot First",
        "Pilot ID",
        "Pilot Member",
        "Business Name",
        "Operator Last",
        "Operator First",
        "Operator ID",
        "Operator Member",
        "A/C Reg",
        "A/C Make",
        "A/C Model",
        "# Passes",
        "Business Address",
        "City",
        "ST",
        "Zip Code",
        "Phone",
        "Email",
    ]
    for i, head in enumerate(header):
        ws.cell(1, i + 1, head)

    for i, file in enumerate(filesToInclude):
        s = SeriesData()
        load_from_db(file=file, s=s)

        # Prefab pilot first/last
        pilot_split = s.info.pilot.rsplit(" ", 1)
        pilot_first = pilot_split[0] if len(pilot_split) >= 1 else ""
        pilot_last = pilot_split[1] if len(pilot_split) == 2 else ""
        # Create an entry for the item
        line_item = [
            s.info.flyin_date,
            s.info.flyin_name,
            s.info.flyin_location,
            s.info.flyin_analyst,
            "Liquid",
            "No",
            pilot_last,
            pilot_first,
            "",
            "",
            s.info.business,
            "",
            "",
            "",
            "",
            s.info.regnum,
            s.info.make,
            s.info.model,
            len(s.passes),
            s.info.street,
            s.info.city,
            s.info.state,
            s.info.zip,
            s.info.phone,
            s.info.email,
        ]

        # Set row to the next blank available. If regnum already exists in sheet, with identical date, then re-assign row to that previous entry's row and totalize passes
        row = ws.max_row + 1
        for i, roww in enumerate(ws.iter_rows()):
            if roww[0].value == line_item[0] and roww[15].value == line_item[15]:
                row = i + 1
                line_item[18] += int(roww[18].value)

        # Write values for line item
        for i, item in enumerate(line_item):
            col = i + 1
            ws.cell(row, col, item)

    # Save it
    wb.save(saveFile)


def export_all_to_excel(series: SeriesData, saveFile: str):
    s = series
    i = s.info

    wb = Workbook()
    # AppInfo Sheet
    ws = wb.active
    ws.title = "App Info"
    # Table Series
    labels = ["series", "created", "modified", "notes_setup", "notes_analyst"]
    vals = [i.series, i.created, i.modified, i.notes_setup, i.notes_analyst]
    # Table Flyin
    labels.extend(["flyin_name", "flyin_location", "flyin_date", "flyin_analyst"])
    vals.extend([i.flyin_name, i.flyin_location, i.flyin_date, i.flyin_analyst])
    # Table Applicator
    labels.extend(
        ["pilot", "business", "street", "city", "state", "zip", "phone", "email"]
    )
    vals.extend(
        [i.pilot, i.business, i.street, i.city, i.state, i.zip, i.phone, i.email]
    )
    # Table Aircraft
    labels.extend(["regnum", "make", "model", "wingspan", "wingspan_units", "winglets"])
    vals.extend([i.regnum, i.make, i.model, i.wingspan, i.wingspan_units, i.winglets])
    # Table Spray System
    labels.extend(
        [
            "target_swath",
            "swath_units",
            "rate",
            "rate_units",
            "pressure",
            "pressure_units",
            "boom_width",
            "boom_width_units",
            "boom_drop",
            "boom_drop_units",
            "nozzle_spacing",
            "nozzle_spacing_units",
        ]
    )
    vals.extend(
        [
            i.swath,
            i.swath_units,
            i.rate,
            i.rate_units,
            i.pressure,
            i.pressure_units,
            i.boom_width,
            i.boom_width_units,
            i.boom_drop,
            i.boom_drop_units,
            i.nozzle_spacing,
            i.nozzle_spacing_units,
        ]
    )
    # Table Nozzles
    for n in i.nozzles:
        labels.extend(
            [
                f"Nozzle {n.id} type",
                f"Nozzle {n.id} size",
                f"Nozzle {n.id} deflection",
                f"Nozzle {n.id} quantity",
            ]
        )
        vals.extend([n.type, n.size, n.deflection, n.quantity])
    # Table Series String
    labels.extend(
        [
            "center_average",
            "smooth_average",
            "equalize_integrals",
            "simulated_adjascent_passes",
        ]
    )
    vals.extend(
        [
            s.string.center_method,
            s.string.smooth,
            s.string.equalize_integrals,
            s.string.simulated_adjascent_passes,
        ]
    )
    # plot it
    for i, (label, val) in enumerate(zip(labels, vals)):
        ws.cell(i + 1, 1, label)
        ws.cell(i + 1, 2, val)
    for cell in ws["A"]:
        cell.style = "Pandas"

    # Passes Sheet
    ws = wb.create_sheet("Passes")
    labels = [
        "pass_name",
        "pass_number",
        "ground_speed",
        "ground_speed_units",
        "spray_height",
        "spray_height_units",
        "pass_heading",
        "wind_direction",
        "wind_speed",
        "wind_speed_units",
        "temperature",
        "temperature_units",
        "humidity",
        "string_include_in_composite",
        "cards_include_in_composite",
        "excitation_wav",
        "emission_wav",
        "trim_left",
        "trim_right",
        "trim_vertical",
        "center",
        "smooth",
        "data_loc_units",
    ]
    for i, label in enumerate(labels):
        ws.cell(i + 1, 1, label)
    p: Pass
    for j, p in enumerate(s.passes):
        vals = [
            p.name,
            p.number,
            p.ground_speed,
            p.ground_speed_units,
            p.spray_height,
            p.spray_height_units,
            p.pass_heading,
            p.wind_direction,
            p.wind_speed,
            p.wind_speed_units,
            p.temperature,
            p.temperature_units,
            p.humidity,
            p.string.include_in_composite,
            p.cards.include_in_composite,
            p.string.trim_l,
            p.string.trim_r,
            p.string.trim_v,
            p.string.center_method,
            p.string.smooth,
            p.string.data_loc_units,
        ]
        for i, val in enumerate(vals):
            ws.cell(i + 1, j + 2, val)
    for cell in ws["A"] + ws[1]:
        cell.style = "Pandas"

    # String Sheet
    ws = wb.create_sheet("String Data")
    # Join all df's
    df = pd.DataFrame()
    for i, p in enumerate(s.passes):
        df = pd.concat([df, p.string.data_ex, p.string.data], axis=1)
    for j, p in enumerate(s.passes):
        ws.cell(1, 1 + (j * 4), p.name)
        ws.merge_cells(
            start_row=1, start_column=1 + (j * 4), end_row=1, end_column=4 + (j * 4)
        )
        ws.cell(2, 1 + (j * 4), "Excitation")
        ws.merge_cells(
            start_row=2, start_column=1 + (j * 4), end_row=2, end_column=2 + (j * 4)
        )
        ws.cell(2, 3 + (j * 4), "Emission")
        ws.merge_cells(
            start_row=2, start_column=3 + (j * 4), end_row=2, end_column=4 + (j * 4)
        )
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1] + ws[2] + ws[3]:
        cell.style = "Pandas"

    # Card Sheet
    ws = wb.create_sheet("Card Data")
    col_pass_name = 2
    for j, p in enumerate(s.passes):
        if p.cards.has_data():
            ws.cell(1, col_pass_name, p.name)
            ws.merge_cells(
                start_row=1,
                start_column=col_pass_name,
                end_row=1,
                end_column=col_pass_name + len(p.cards.card_list) - 1,
            )
            for k, c in enumerate(p.cards.card_list):
                c: SprayCard
                labels = [
                    "name",
                    "has_image",
                    "include_in_composite",
                    "location",
                    "units",
                    "threshold_type",
                    "threshold_method_grayscale",
                    "threshold_grayscale",
                    "threshold_color_hue_min",
                    "threshold_color_hue_max",
                    "threshold_color_hue_pass",
                    "threshold_color_saturation_min",
                    "threshold_color_saturation_max",
                    "threshold_color_saturation_pass",
                    "threshold_color_brightness_min",
                    "threshold_color_brightness_max",
                    "threshold_color_brightness_pass",
                    "dpi",
                    "spread_method",
                    "sf_a",
                    "sf_b",
                    "sf_c",
                ]
                vals = [
                    c.name,
                    c.has_image,
                    c.include_in_composite,
                    c.location,
                    c.location_units,
                    c.threshold_type,
                    c.threshold_method_grayscale,
                    c.threshold_grayscale,
                    c.threshold_color_hue_min,
                    c.threshold_color_hue_max,
                    c.threshold_color_hue_pass,
                    c.threshold_color_saturation_min,
                    c.threshold_color_saturation_max,
                    c.threshold_color_saturation_pass,
                    c.threshold_color_brightness_min,
                    c.threshold_color_brightness_max,
                    c.threshold_color_brightness_pass,
                    c.dpi,
                    c.spread_method,
                    c.spread_factor_a,
                    c.spread_factor_b,
                    c.spread_factor_c,
                ]
                for i, (label, val) in enumerate(zip(labels, vals)):
                    ws.cell(2 + i, 1, label)
                    ws.cell(2 + i, col_pass_name + k, val)
            col_pass_name += len(p.cards.card_list)

    for cell in ws["A"] + ws[1] + ws[2]:
        cell.style = "Pandas"

    # Save it
    wb.save(saveFile)
