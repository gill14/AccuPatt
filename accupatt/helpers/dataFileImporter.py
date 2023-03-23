import io
import os

import accupatt.config as cfg
import numpy as np
import openpyxl
import pandas as pd
from accupatt.helpers.dBBridge import load_from_db, save_to_db
from accupatt.models.appInfo import AppInfo, Nozzle
from accupatt.models.passData import Pass
from accupatt.models.seriesData import SeriesData
from accupatt.models.sprayCard import SprayCard
from openpyxl_image_loader import SheetImageLoader


def get_file_type(file) -> int:
    if len(file) > 3 and file[-2:] == "db":
        return cfg.DATA_FILE_TYPE_ACCUPATT
    elif len(file) > 3 and file[-4:] == "xlsx":
        return cfg.DATA_FILE_TYPE_ACCUPATT_LEGACY
    elif len(file) > 3 and file[-1] == "A":
        return cfg.DATA_FILE_TYPE_WRK
    elif len(file) > 6 and file[-4:] == ".txt":
        return cfg.DATA_FILE_TYPE_USDA
    else:
        return cfg.DATA_FILE_TYPE_NONE


def save_file(file, series: SeriesData) -> bool:
    if get_file_type(file) != cfg.DATA_FILE_TYPE_ACCUPATT:
        return False
    if save_to_db(file, series):
        return True
    else:
        return False


def load_file_to_series(file, series: SeriesData, info_only=False):
    match get_file_type(file):
        case cfg.DATA_FILE_TYPE_ACCUPATT:
            load_from_db(file, series, info_only)
        case cfg.DATA_FILE_TYPE_ACCUPATT_LEGACY:
            load_from_accupatt_1_file(file, series)
        case cfg.DATA_FILE_TYPE_USDA:
            load_from_usda_file(file, series)
        case cfg.DATA_FILE_TYPE_WRK:
            load_from_wrk_file(file, series)
        case _:
            # Unknown file type or none
            pass


def convert_import_to_db(file, prog=None) -> str:
    if prog is not None:
        prog.setRange(0, 3)
        prog.setValue(0)
        prog.setLabelText("Loading in Series Data")
    s = SeriesData()
    load_file_to_series(file, s)
    if prog is not None:
        prog.setLabelText("Creating Local Database")
        prog.setValue(1)
    # Write to DB (same dir as original xlsx)
    file_db = os.path.join(os.path.dirname(file), s.info.string_reg_series() + ".db")
    save_to_db(file=file_db, s=s)
    if get_file_type(file) == cfg.DATA_FILE_TYPE_ACCUPATT_LEGACY:
        # Need to check for card images
        if prog is not None:
            prog.setLabelText("Checking XLSX for Spray Cards")
            prog.setValue(2)
        wb = openpyxl.load_workbook(file)
        if "Card Data" in wb.sheetnames:
            sh = wb["Card Data"]
            on_pass = int(sh["B2"].value)
            # Loop over declard cards and save images to db if has_image
            p: Pass = s.passes[on_pass - 1]
            c: SprayCard
            for i, c in enumerate(p.cards.card_list):
                if prog is not None:
                    if i == 0:
                        prog.setRange(0, len(p.cards.card_list))
                    prog.setValue(i)
                    prog.setLabelText(f"Copying image for {c.name}")
                c.filepath = file_db
                if c.has_image:
                    # Get the image from applicable sheet
                    image_loader = SheetImageLoader(wb[c.name])
                    image = image_loader.get("A1")
                    # Conver to bytestream
                    stream = io.BytesIO()
                    image.save(stream, format="PNG")
                    # Save it to the database
                    c.save_image_to_file(stream.getvalue())
                    # Reclaim resources
                    stream.close()
                if i == len(p.cards.card_list) - 1:
                    prog.setValue(i + 1)
    if prog is not None:
        prog.setValue(3)
    return file_db


def load_from_accupatt_1_file(file, s: SeriesData):
    # indicator for metric
    isMetric = False

    # Load entire WB into dict of sheets
    df_map = pd.read_excel(file, sheet_name=None, header=None)

    i = s.info
    # Pull data from Fly-In Tab
    df = df_map["Fly-In Data"].fillna("")
    i.flyin_name = df.iat[0, 0]
    i.flyin_location = df.iat[1, 0]
    i.flyin_date = df.iat[2, 0]
    i.flyin_analyst = df.iat[3, 0]

    # Pull data from AppInfo Tab
    df = df_map["Aircraft Data"].fillna("")
    i.pilot = df.iat[0, 1]
    i.business = df.iat[1, 1]
    i.phone = df.iat[2, 1]
    i.street = df.iat[3, 1]
    i.city = df.iat[4, 1]
    i.state = df.iat[5, 1]
    i.regnum = df.iat[6, 1]
    i.series = df.iat[7, 1]
    i.make = df.iat[8, 1]
    i.model = df.iat[9, 1]
    if (noz_type_1 := df.iat[10, 1]) != "":
        i.nozzles.append(
            translateNozzle(
                id=1,
                type=noz_type_1,
                size=df.iat[12, 1],
                defl=df.iat[13, 1],
                quant=df.iat[11, 1],
            )
        )

    if (noz_type_2 := df.iat[14, 1]) != "":
        i.nozzles.append(
            translateNozzle(
                id=2,
                type=noz_type_2,
                size=df.iat[16, 1],
                defl=df.iat[17, 1],
                quant=df.iat[15, 1],
            )
        )
    i.set_pressure(df.iat[18, 1])
    i.set_rate(df.iat[19, 1])
    i.set_swath(df.iat[20, 1])
    s.string.set_swath_adjusted(df.iat[20, 1])  # Just in case it isn't set below
    s.cards.set_swath_adjusted(df.iat[20, 1])  # Just in case it isn't set below
    i.set_wingspan(df.iat[27, 1])
    i.set_boom_width(df.iat[28, 1])
    i.set_boom_drop(df.iat[30, 1])
    i.set_nozzle_spacing(df.iat[31, 1])
    i.winglets = df.iat[32, 1]
    i.notes_setup = df.iat[33, 1]
    # Col 2 if available
    if df.shape[1] > 2:
        i.email = df.iat[0, 2]
        i.zip = df.iat[5, 2]
        if i.zip != "":
            i.zip = str(int(i.zip))
        s.string.set_swath_adjusted(df.iat[20, 2])  # Just in case it isn't set below
    s.cards.set_swath_adjusted(df.iat[20, 2])
    # Set units for series/passes based on 'metric' identifier
    isMetric = df.iat[35, 1] == "TRUE"
    s.info.swath_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.rate_units = cfg.UNIT_LPHA if isMetric else cfg.UNIT_GPA
    s.info.pressure_units = cfg.UNIT_BAR if isMetric else cfg.UNIT_PSI
    s.info.wingspan_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.boom_width_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.boom_drop_units = cfg.UNIT_CM if isMetric else cfg.UNIT_IN
    s.info.nozzle_spacing_units = cfg.UNIT_CM if isMetric else cfg.UNIT_IN

    # Pull data from Series Data tab
    df = df_map["Series Data"].fillna("")
    df.columns = df.iloc[0]
    # Search for any active passes and create entries in seriesData.passes dict
    for i, column in enumerate(df.columns[1:]):
        if str(df.at[1, column]) != "":
            p = Pass(number=i + 1)
            p.set_ground_speed(
                df.at[1, column], units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH
            )
            p.set_spray_height(
                df.at[2, column], units=cfg.UNIT_M if isMetric else cfg.UNIT_FT
            )
            p.set_pass_heading(df.at[3, column])
            p.set_wind_direction(df.at[4, column])
            p.set_wind_speed(
                df.at[5, column], units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH
            )
            p.set_temperature(
                df.at[6, "Pass 1"], units=cfg.UNIT_DEG_C if isMetric else cfg.UNIT_DEG_F
            )
            p.set_humidity(df.at[7, "Pass 1"])
            p.string.data_loc_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
            s.passes.append(p)

    # Pull data from Pattern Data tab
    df: pd.DataFrame = df_map["Pattern Data"].fillna(np.nan)
    if df.shape[1] < 13:
        df["nan"] = np.nan
    # Make new empty dataframe for info
    # df_info = pd.DataFrame({'Pass 1':[], 'Pass 2':[], 'Pass 3':[], 'Pass 4':[], 'Pass 5':[], 'Pass 6':[]})
    cols = ["loc", "Pass 1", "Pass 2", "Pass 3", "Pass 4", "Pass 5", "Pass 6"]
    trims = df.iloc[0:3, [2, 4, 6, 8, 10, 12]].reset_index(drop=True).astype(float)
    trims.columns = cols[1:]
    params1 = pd.DataFrame(
        df.iloc[3:4, [1, 3, 5, 7, 9, 11]].reset_index(drop=True), columns=cols[1:]
    )
    params2 = pd.DataFrame(
        df.iloc[4:4, [2, 4, 6, 8, 10, 12]].reset_index(drop=True), columns=cols[1:]
    )
    params = pd.concat([params1, params2])
    df_em = df.iloc[5:, [0, 2, 4, 6, 8, 10, 12]].reset_index(drop=True).astype(float)
    df_em.columns = cols
    df_ex = df.iloc[5:, [0, 1, 3, 5, 7, 9, 11]].reset_index(drop=True).astype(float)
    df_ex.columns = cols

    # Pull patterns and place them into seriesData.passes list by name (created above)
    for p in s.passes:
        p.string.trim_l = (
            0 if np.isnan(trims.at[0, p.name]) else int(trims.at[0, p.name])
        )
        p.string.trim_r = (
            0 if np.isnan(trims.at[1, p.name]) else int(trims.at[1, p.name])
        )
        p.string.trim_v = 0 if np.isnan(trims.at[2, p.name]) else trims.at[2, p.name]
        # TODO Integration Time is params row 0, must convert to int
        # TODO Ex/Em Wavs are params rows 1, 2 respective, must strip string
        p.string.data = df_em[["loc", p.name]]
        p.string.data_ex = df_ex[["loc", p.name]]

    # Create SprayCards if applicable
    if "Card Data" in df_map.keys():
        wb = openpyxl.load_workbook(file, read_only=True)
        sh = wb["Card Data"]
        spacing = sh["B1"].value
        on_pass = int(sh["B2"].value)
        spread_method_text = sh["B3"].value
        spread_method = cfg.SPREAD_METHOD_ADAPTIVE
        if spread_method_text == cfg.SPREAD_METHOD_DIRECT:
            spread_method = cfg.SPREAD_METHOD_DIRECT
        if spread_method_text == cfg.SPREAD_METHOD_NONE:
            spread_method = cfg.SPREAD_METHOD_NONE
        spread_a = sh["B4"].value
        spread_b = sh["B5"].value
        spread_c = sh["B6"].value
        # Card Pass
        p: Pass = s.passes[on_pass - 1]
        for col in range(5, 14):
            if not sh.cell(row=1, column=col).value:
                continue
            c = SprayCard(name=sh.cell(row=1, column=col).value)
            c.filepath = file
            c.location = (col - 9) * spacing
            c.location_units = cfg.UNIT_FT
            c.has_image = 1 if sh.cell(row=2, column=col).value else 0
            c.include_in_composite = 1 if sh.cell(row=3, column=col).value else 0
            if c.has_image:
                c.threshold_grayscale = int(sh.cell(row=4, column=col).value)
            c.dpi = 600
            c.threshold_type = cfg.THRESHOLD_TYPE_GRAYSCALE
            c.spread_method = spread_method
            c.spread_factor_a = spread_a
            c.spread_factor_b = spread_b
            c.spread_factor_c = spread_c
            p.cards.card_list.append(c)


def translateNozzle(id, type, size, defl, quant) -> Nozzle:
    conv = {
        "CP11TT 20Deg FF": "CP11TT 20°FF",
        "CP11TT 40Deg FF": "CP11TT 40°FF",
        "CP11TT 80Deg FF": "CP11TT 80°FF",
        "Hollow Cone Steel DC45": "Steel Disc Core 45",
        "Hollow Cone Ceramic DC45": "Ceramic Disc Core 45",
        "40Deg FF": "Standard 40°FF",
        "80Deg FF": "Standard 80°FF",
    }
    if type in conv.keys():
        type = conv[type]
    try:
        f = float(size)
        if f.is_integer():
            size = int(f)
    except:
        pass
    try:
        f = float(defl)
        if f.is_integer():
            defl = int(f)
    except:
        pass
    try:
        i = int(quant)
        quant = i
    except:
        pass
    return Nozzle(id, type, str(size), str(defl), quant)


def load_image_from_accupatt_1(file, spray_card_name):
    wb = openpyxl.load_workbook(file)
    # Get (PIL) Image from applicable sheet
    return SheetImageLoader(wb[spray_card_name]).get("A1")


def load_from_accustain_file(file, s: SeriesData):
    # Load entire WB into dict of sheets
    df_map = pd.read_excel(file, sheet_name=None, header=None)

    df_settings = df_map["Settings"]

    df_appinfo = df_map["AppInfo"].fillna("")
    s.info.set_swath(df_appinfo.iat[3, 6])
    s.cards.set_swath_adjusted(df_appinfo.iat[3, 6])

    df_index = df_map["Index"]
    # TODO


def load_from_usda_file(file: str, s: SeriesData):
    # Split file name for parts
    parts = file.split(os.sep)[-1].split(" ")
    print(file.split(os.sep))
    print(parts)
    regnum = parts[0]
    print(regnum)
    series_letter = parts[1]
    print(series_letter)
    id = regnum + " " + series_letter

    # Get a sorted list of pass files for this series
    dir = os.path.dirname(file)
    files = [
        os.path.join(dir, f)
        for f in os.listdir(dir)
        if os.path.isfile(os.path.join(dir, f))
    ]
    files = [fn for fn in files if id in fn]
    files.sort()

    i = s.info

    i.regnum = regnum
    i.series = ord(series_letter.lower()) - 96
    print(i.series)
    i.swath = 65  # Hard default added to analyst notes below

    # get pilot parameters file
    if os.path.isfile(os.path.join(dir, "Pilot Paramters.prn")):
        pdf = pd.read_csv(
            os.path.join(dir, "Pilot Paramters.prn"), sep="\t", index_col=0
        )
        i.pilot = pdf.loc[regnum, "Pilot Name"]
        i.street = pdf.loc[regnum, "Street Address"]
        i.city = pdf.loc[regnum, "City"]
        i.state = pdf.loc[regnum, "State"]
        i.zip = pdf.loc[regnum, "Zip Code"]
        # put the rest in notes because it was unreliably used
        note0 = "These fields listed for reference only from original data file."
        note1 = f"Applicator License #: {pdf.loc[regnum,'Applicator Licence Number']}"
        note2 = f"Office #: {pdf.loc[regnum, 'Office Number']}"
        note3 = f"Cell #: {pdf.loc[regnum, 'Cell Number']}"
        note4 = f"Email: {pdf.loc[regnum, 'Email']}"
        note5 = "SWATH WIDTH SET TO 65 FT BY GLOBAL DEFAULT FOR USDA FILES"
        i.notes_analyst = "\r".join([note0, note1, note2, note3, note4, note5])

    for file in files:
        p = Pass(number=int(parts[2]))

        lines = []
        with open(file) as ffile:
            lines = ffile.readlines()
        d_ex = []
        d_em = []
        for i, line in enumerate(lines):
            if i < 2:
                continue
            line_item = line.strip().split("\t")
            d_ex.append({"loc": float(line_item[1]), p.name: float(line_item[3])})
            d_em.append({"loc": float(line_item[1]), p.name: float(line_item[2])})
        p.string.data_ex = pd.DataFrame(d_ex)
        p.string.data = pd.DataFrame(d_em)

        s.passes.append(p)


def load_from_wrk_file(file, s: SeriesData):
    isMetric = False  # This is a bold assumption

    # Read in file to a list of lines, stripping line return and quotes
    lines = []
    with open(file) as ffile:
        lines = ffile.readlines()
    for i, line in enumerate(lines):
        lines[i] = lines[i].strip()
        lines[i] = lines[i].strip('"')
        print("Line {}: {}".format(i, lines[i]))

    i = s.info
    i.regnum = file.split(os.sep)[-1][2:-3]
    i.series = int(file[-2])

    i.flyin_name = lines[0]
    i.flyin_location = lines[1]
    i.flyin_date = lines[2]
    i.flyin_analyst = lines[3]

    _fl_length = lines[4]
    _analysis_speed = lines[5]
    _n1flow40 = lines[6]
    _n2flow40 = lines[7]
    _n2q = int(lines[8].strip() or 0)
    _n1s = lines[9]
    _n2s = lines[10]
    _n1s_name = lines[11]
    _n2s_name = lines[12]
    _n1d = lines[13]
    _n2d = lines[14]

    i.business = lines[15]
    i.street = lines[16]
    i.city = lines[17]
    i.state = lines[18]
    i.zip = lines[19]
    i.phone = lines[20]
    i.pilot = lines[21]
    # i.regnum = lines[22]
    i.model = lines[23]

    _n1t = lines[24]
    _n2t = (
        _n1t  # WRK only allowed selection of different orifice sizes, not nozzle types
    )
    _n_total_q = int(lines[25].strip() or 0)
    _n1q = _n_total_q - _n2q
    if _n1q > 0:
        i.nozzles.append(
            translateNozzle(id=1, type=_n1t, size=_n1s, defl=_n1d, quant=_n1q)
        )
    if _n2q > 0:
        i.nozzles.append(
            translateNozzle(id=2, type=_n2t, size=_n2s, defl=_n2d, quant=_n2q)
        )

    i.pressure = int(lines[26])
    i.rate = float(lines[27])
    i.swath = int(lines[28])

    # Get a sorted list of pass files for this series
    dir = os.path.dirname(file)
    files = [f for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))]
    files = [os.path.join(dir, fn) for fn in files if fn[:-1] in file]
    files.sort()

    for file in files:
        lines = []
        with open(file) as ffile:
            lines = ffile.readlines()
        for i, line in enumerate(lines):
            lines[i] = lines[i].strip()
            lines[i] = lines[i].strip('"')

        p = Pass(number=ord(file[-1].lower()) - 96)
        p.set_ground_speed(
            int(lines[29]), units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH
        )
        p.set_wind_speed(
            float(lines[30]), units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH
        )
        p.set_wind_direction(int(lines[31]))
        # Complicated Pass Heading conversion:
        _cw = float(lines[32])
        """if _cw > 0:
            _ph = (180/math.pi) *(math.asin(_cw) / float(lines[30])) + float(lines[31])
        else:
            _ph = abs((180/math.pi)*(math.asin(abs(_cw)) / float(lines[30])) - float(lines[31]))"""
        _ph = 0  # TODO
        p.pass_heading = _ph
        p.set_temperature(
            int(lines[33]), units=cfg.UNIT_DEG_C if isMetric else cfg.UNIT_DEG_F
        )
        p.set_spray_height(
            float(lines[34]), units=cfg.UNIT_M if isMetric else cfg.UNIT_FT
        )
        p.set_humidity(int(lines[35]))

        # calculate data point spacing from fl length and num data points
        _fl = int(lines[4])
        _num_data_points = int(lines[37])
        _spacing = _fl / _num_data_points

        # loop over data points and get them into a dataframe
        d = []
        for i in range(38, 38 + _num_data_points):
            d.append({"loc": (i - 38) * _spacing, p.name: float(lines[i] or 0)})
        p.string.data = pd.DataFrame(d)

        # Turn off smoothing
        p.string.smooth = False

        s.passes.append(p)

    s.string.smooth = False
