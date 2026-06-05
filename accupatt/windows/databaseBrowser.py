from __future__ import annotations

import os
import sqlite3
import subprocess
from dataclasses import dataclass
from datetime import date, datetime
from enum import Enum
from io import StringIO
from typing import Optional, Union

import accupatt.config as cfg
import pandas as pd
from aerial_spray_nozzle_models import AtomizationModelMulti
from openpyxl import Workbook
from PyQt6.QtCore import (
    QAbstractTableModel,
    QDate,
    QModelIndex,
    QObject,
    QPoint,
    QSortFilterProxyModel,
    QThread,
    Qt,
    pyqtSignal,
    pyqtSlot,
)
from PyQt6.QtGui import QAction, QBrush, QColor
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QTableView,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from aerial_spray_nozzle_models.reference import DSC_REFERENCE

from accupatt.models.passData import Pass
from accupatt.models.seriesDataString import SeriesDataString


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_with_units(value, units) -> str:
    v = str(value or "").strip()
    u = str(units or "").strip()
    if not v:
        return ""
    return f"{v} {u}".strip() if u else v


def _fmt_timestamp(ts) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OSError):
        return str(ts)


def _parse_leading_float(value: str) -> Optional[float]:
    parts = value.strip().split()
    if not parts:
        return None
    try:
        return float(parts[0])
    except ValueError:
        return None


def _parse_date(value: str) -> Optional[date]:
    for fmt in ("%Y-%m-%d", "%d %b %Y"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _to_mph(value, units: str) -> Optional[float]:
    try:
        v = float(value)
        u = str(units or "").strip().lower()
        if u == cfg.UNIT_MPH:
            return v
        if u == cfg.UNIT_KPH:
            return v * cfg.MPH_PER_KPH
        if u == cfg.UNIT_KN:
            return v * cfg.MPH_PER_KN
    except (TypeError, ValueError):
        pass
    return None


def _to_psi(value, units: str) -> Optional[float]:
    try:
        v = float(value)
        u = str(units or "").strip().lower()
        if u == cfg.UNIT_PSI:
            return v
        if u == cfg.UNIT_BAR:
            return v * cfg.PSI_PER_BAR
        if u == cfg.UNIT_KPA:
            return v / cfg.KPA_PER_PSI
    except (TypeError, ValueError):
        pass
    return None


# ---------------------------------------------------------------------------
# Filter types and state
# ---------------------------------------------------------------------------

class FilterType(Enum):
    TEXT = "text"
    NUMERIC = "numeric"
    DATE = "date"


@dataclass
class ColumnFilter:
    filter_type: FilterType
    text: str = ""
    num_min: Optional[float] = None
    num_max: Optional[float] = None
    date_min: Optional[date] = None
    date_max: Optional[date] = None

    def is_active(self) -> bool:
        if self.filter_type == FilterType.TEXT:
            return bool(self.text.strip())
        if self.filter_type == FilterType.NUMERIC:
            return self.num_min is not None or self.num_max is not None
        if self.filter_type == FilterType.DATE:
            return self.date_min is not None or self.date_max is not None
        return False

    def matches(self, value: str) -> bool:
        if self.filter_type == FilterType.TEXT:
            return self.text.strip().lower() in value.lower()

        if self.filter_type == FilterType.NUMERIC:
            num = _parse_leading_float(value)
            if num is None:
                return self.num_min is None and self.num_max is None
            if self.num_min is not None and num < self.num_min:
                return False
            if self.num_max is not None and num > self.num_max:
                return False
            return True

        if self.filter_type == FilterType.DATE:
            d = _parse_date(value)
            if d is None:
                return self.date_min is None and self.date_max is None
            if self.date_min is not None and d < self.date_min:
                return False
            if self.date_max is not None and d > self.date_max:
                return False
            return True

        return True


# ---------------------------------------------------------------------------
# Filter dialogs
# ---------------------------------------------------------------------------

class _FilterDialogBase(QDialog):
    def __init__(self, col_label: str, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle(f"Filter: {col_label}")
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.WindowCloseButtonHint)
        self.setModal(True)
        self._result: Optional[ColumnFilter] = None

    def result_filter(self) -> Optional[ColumnFilter]:
        return self._result

    def _make_button_box(self) -> QDialogButtonBox:
        box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        self._btn_clear = box.addButton(
            "Clear Filter", QDialogButtonBox.ButtonRole.ResetRole
        )
        box.accepted.connect(self._on_ok)
        box.rejected.connect(self.reject)
        self._btn_clear.clicked.connect(self._on_clear)
        return box

    def _on_ok(self):
        raise NotImplementedError

    def _on_clear(self):
        self._result = None
        self.accept()


class TextFilterDialog(_FilterDialogBase):
    def __init__(self, col_label: str, existing: Optional[ColumnFilter] = None,
                 parent: Optional[QWidget] = None):
        super().__init__(col_label, parent)
        self.setFixedWidth(320)
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        self._input = QLineEdit()
        self._input.setPlaceholderText("leave blank to match all")
        self._input.returnPressed.connect(self._on_ok)
        if existing and existing.is_active():
            self._input.setText(existing.text)
        self._input.selectAll()
        layout.addWidget(QLabel("Contains:"))
        layout.addWidget(self._input)
        layout.addWidget(self._make_button_box())

    def showEvent(self, event):
        super().showEvent(event)
        self._input.setFocus()

    def _on_ok(self):
        text = self._input.text().strip()
        self._result = ColumnFilter(FilterType.TEXT, text=text) if text else None
        self.accept()


class NumericRangeDialog(_FilterDialogBase):
    def __init__(self, col_label: str, existing: Optional[ColumnFilter] = None,
                 parent: Optional[QWidget] = None):
        super().__init__(col_label, parent)
        self.setFixedWidth(280)
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        grid = QGridLayout()
        grid.setSpacing(6)
        self._min_input = QLineEdit()
        self._min_input.setPlaceholderText("no limit")
        self._max_input = QLineEdit()
        self._max_input.setPlaceholderText("no limit")
        if existing and existing.is_active():
            if existing.num_min is not None:
                self._min_input.setText(str(existing.num_min))
            if existing.num_max is not None:
                self._max_input.setText(str(existing.num_max))
        grid.addWidget(QLabel("Min:"), 0, 0)
        grid.addWidget(self._min_input, 0, 1)
        grid.addWidget(QLabel("Max:"), 1, 0)
        grid.addWidget(self._max_input, 1, 1)
        note = QLabel("Leave blank for no limit.")
        note.setStyleSheet("color: gray; font-size: 11px;")
        layout.addLayout(grid)
        layout.addWidget(note)
        layout.addWidget(self._make_button_box())

    def showEvent(self, event):
        super().showEvent(event)
        self._min_input.setFocus()

    def _on_ok(self):
        num_min: Optional[float] = None
        num_max: Optional[float] = None
        min_text = self._min_input.text().strip()
        max_text = self._max_input.text().strip()
        if min_text:
            try:
                num_min = float(min_text)
            except ValueError:
                QMessageBox.warning(self, "Invalid Input", f"'{min_text}' is not a valid number.")
                self._min_input.setFocus(); self._min_input.selectAll()
                return
        if max_text:
            try:
                num_max = float(max_text)
            except ValueError:
                QMessageBox.warning(self, "Invalid Input", f"'{max_text}' is not a valid number.")
                self._max_input.setFocus(); self._max_input.selectAll()
                return
        if num_min is not None and num_max is not None and num_min > num_max:
            QMessageBox.warning(self, "Invalid Range", "Min cannot be greater than Max.")
            return
        self._result = (
            ColumnFilter(FilterType.NUMERIC, num_min=num_min, num_max=num_max)
            if (num_min is not None or num_max is not None)
            else None
        )
        self.accept()


class DateRangeDialog(_FilterDialogBase):
    def __init__(self, col_label: str, existing: Optional[ColumnFilter] = None,
                 parent: Optional[QWidget] = None):
        super().__init__(col_label, parent)
        self.setFixedWidth(300)
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        today = QDate.currentDate()
        self._from_check = QCheckBox("From:")
        self._from_edit = QDateEdit()
        self._from_edit.setCalendarPopup(True)
        self._from_edit.setDisplayFormat("yyyy-MM-dd")
        self._to_check = QCheckBox("To:")
        self._to_edit = QDateEdit()
        self._to_edit.setCalendarPopup(True)
        self._to_edit.setDisplayFormat("yyyy-MM-dd")
        if existing and existing.is_active():
            if existing.date_min is not None:
                self._from_check.setChecked(True)
                self._from_edit.setDate(QDate(existing.date_min.year, existing.date_min.month, existing.date_min.day))
            else:
                self._from_check.setChecked(False)
                self._from_edit.setDate(today)
            if existing.date_max is not None:
                self._to_check.setChecked(True)
                self._to_edit.setDate(QDate(existing.date_max.year, existing.date_max.month, existing.date_max.day))
            else:
                self._to_check.setChecked(False)
                self._to_edit.setDate(today)
        else:
            self._from_check.setChecked(False)
            self._from_edit.setDate(today)
            self._to_check.setChecked(False)
            self._to_edit.setDate(today)
        self._from_edit.setEnabled(self._from_check.isChecked())
        self._to_edit.setEnabled(self._to_check.isChecked())
        self._from_check.toggled.connect(self._from_edit.setEnabled)
        self._to_check.toggled.connect(self._to_edit.setEnabled)
        grid = QGridLayout()
        grid.setSpacing(6)
        grid.addWidget(self._from_check, 0, 0)
        grid.addWidget(self._from_edit, 0, 1)
        grid.addWidget(self._to_check, 1, 0)
        grid.addWidget(self._to_edit, 1, 1)
        layout.addLayout(grid)
        layout.addWidget(self._make_button_box())

    def _on_ok(self):
        date_min: Optional[date] = None
        date_max: Optional[date] = None
        if self._from_check.isChecked():
            qd = self._from_edit.date()
            date_min = date(qd.year(), qd.month(), qd.day())
        if self._to_check.isChecked():
            qd = self._to_edit.date()
            date_max = date(qd.year(), qd.month(), qd.day())
        if date_min is not None and date_max is not None and date_min > date_max:
            QMessageBox.warning(self, "Invalid Range", "From date cannot be after To date.")
            return
        self._result = (
            ColumnFilter(FilterType.DATE, date_min=date_min, date_max=date_max)
            if (date_min is not None or date_max is not None)
            else None
        )
        self.accept()


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SeriesRecord:
    filepath: str = ""
    filename: str = ""
    series_id: str = ""
    # Fly-In
    flyin_name: str = ""
    flyin_location: str = ""
    flyin_date: str = ""
    flyin_analyst: str = ""
    # Applicator
    pilot: str = ""
    business: str = ""
    street: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    phone: str = ""
    email: str = ""
    # Aircraft
    regnum: str = ""
    series_num: str = ""
    make: str = ""
    model: str = ""
    wingspan: str = ""
    winglets: str = ""
    # Spray System
    swath: str = ""
    rate: str = ""
    pressure: str = ""
    boom_drop: str = ""
    nozzle_spacing: str = ""
    # Nozzle Set 1
    nozzle1_type: str = ""
    nozzle1_size: str = ""
    nozzle1_deflection: str = ""
    nozzle1_quantity: str = ""
    # Observables
    avg_speed: str = ""
    avg_height: str = ""
    pass_count: str = ""
    notes_setup: str = ""
    notes_analyst: str = ""
    # Analysis
    string_swath_adjusted: str = ""
    card_swath_adjusted: str = ""
    rt_cv: str = ""
    bf_cv: str = ""
    modeled_dsc: str = ""
    # Internal (not shown as columns)
    created: str = ""
    modified: str = ""


@dataclass
class ColumnDef:
    key: str
    label: str
    default_visible: bool
    alignment: Qt.AlignmentFlag = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
    filter_type: FilterType = FilterType.TEXT


@dataclass
class ColumnSection:
    """Sentinel inserted into ALL_COLUMNS to add a group header in the Columns menu."""
    title: str


_LEFT = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
_CENTER = Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
_RIGHT = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
_T = FilterType.TEXT
_N = FilterType.NUMERIC
_D = FilterType.DATE

# Flat list of ColumnDef items interleaved with ColumnSection sentinels.
# ColumnSection items are used only by the Columns dropdown; the table model
# filters them out.
ALL_COLUMNS: list[Union[ColumnDef, ColumnSection]] = [
    ColumnDef("filename",             "File",                False, _LEFT,   _T),
    ColumnSection("Fly-In"),
    ColumnDef("flyin_name",           "Event Name",          True,  _LEFT,   _T),
    ColumnDef("flyin_location",       "Location",            False, _LEFT,   _T),
    ColumnDef("flyin_date",           "Date",                True,  _CENTER, _D),
    ColumnDef("flyin_analyst",        "Analyst(s)",          False, _LEFT,   _T),
    ColumnSection("Applicator"),
    ColumnDef("pilot",                "Pilot",               True,  _LEFT,   _T),
    ColumnDef("business",             "Business",            True,  _LEFT,   _T),
    ColumnDef("street",               "Street",              False, _LEFT,   _T),
    ColumnDef("city",                 "City",                True,  _LEFT,   _T),
    ColumnDef("state",                "State",               True,  _LEFT,   _T),
    ColumnDef("zip_code",             "ZIP",                 False, _LEFT,   _T),
    ColumnDef("phone",                "Phone",               False, _LEFT,   _T),
    ColumnDef("email",                "Email",               False, _LEFT,   _T),
    ColumnSection("Aircraft"),
    ColumnDef("regnum",               "Reg #",               True,  _CENTER, _T),
    ColumnDef("series_num",           "Series #",            True,  _CENTER, _N),
    ColumnDef("make",                 "Make",                True,  _LEFT,   _T),
    ColumnDef("model",                "Model",               True,  _LEFT,   _T),
    ColumnDef("wingspan",             "Wingspan",            False, _RIGHT,  _N),
    ColumnDef("winglets",             "Winglets",            False, _LEFT,   _T),
    ColumnSection("Spray System"),
    ColumnDef("swath",                "Target Swath",        True,  _RIGHT,  _N),
    ColumnDef("rate",                 "Target Rate",         True,  _RIGHT,  _N),
    ColumnDef("pressure",             "Boom Pressure",       False, _RIGHT,  _N),
    ColumnDef("boom_drop",            "Boom Drop",           False, _RIGHT,  _N),
    ColumnDef("nozzle_spacing",       "Nozzle Spacing",      False, _RIGHT,  _N),
    ColumnSection("Nozzle Set 1"),
    ColumnDef("nozzle1_type",         "Type",                True,  _LEFT,   _T),
    ColumnDef("nozzle1_size",         "Size",                True,  _CENTER, _N),
    ColumnDef("nozzle1_deflection",   "Deflection",          True,  _CENTER, _N),
    ColumnDef("nozzle1_quantity",     "Quantity",            True,  _CENTER, _N),
    ColumnSection("Observables"),
    ColumnDef("avg_speed",            "Avg Airspeed",        False, _RIGHT,  _N),
    ColumnDef("avg_height",           "Avg Height",          False, _RIGHT,  _N),
    ColumnDef("pass_count",           "# Passes",            False, _CENTER, _N),
    ColumnDef("notes_setup",          "Notes (Setup)",       False, _LEFT,   _T),
    ColumnDef("notes_analyst",        "Notes (Analyst)",     False, _LEFT,   _T),
    ColumnSection("Analysis"),
    ColumnDef("string_swath_adjusted","Adj. Swath (String)", True,  _RIGHT,  _N),
    ColumnDef("card_swath_adjusted",  "Adj. Swath (Card)",   False, _RIGHT,  _N),
    ColumnDef("rt_cv",                "RT CV",               True,  _RIGHT,  _N),
    ColumnDef("bf_cv",                "B&F CV",              True,  _RIGHT,  _N),
    ColumnDef("modeled_dsc",          "Modeled DSC",         True,  _CENTER, _T),
]

# Filtered view — only ColumnDef items, in order
_COLUMN_DEFS: list[ColumnDef] = [c for c in ALL_COLUMNS if isinstance(c, ColumnDef)]
_COLUMN_DEFS_BY_KEY: dict[str, ColumnDef] = {c.key: c for c in _COLUMN_DEFS}

# Maps DSC category string → QBrush (alpha 128, matching card analysis table)
def _make_dsc_brushes() -> dict[str, QBrush]:
    out = {}
    for cat, entry in DSC_REFERENCE.items():
        c = QColor(entry["Color"])
        c.setAlpha(128)
        out[cat] = QBrush(c)
    return out

_DSC_BRUSHES: dict[str, QBrush] = _make_dsc_brushes()


# ---------------------------------------------------------------------------
# Background scanner
# ---------------------------------------------------------------------------

class DatabaseScanner(QObject):
    progress = pyqtSignal(int, int)
    record_found = pyqtSignal(object)
    finished = pyqtSignal(int)
    scan_error = pyqtSignal(str, str)

    def __init__(self, directory: str):
        super().__init__()
        self.directory = directory
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    @pyqtSlot()
    def run(self):
        db_files: list[str] = []
        for root, _, files in os.walk(self.directory):
            for f in files:
                if f.lower().endswith(".db"):
                    db_files.append(os.path.join(root, f))
        db_files.sort()

        total = len(db_files)
        found = 0
        for i, filepath in enumerate(db_files):
            if self._cancelled:
                break
            self.progress.emit(i + 1, total)
            try:
                record = self._read_record(filepath)
                if record is not None:
                    self.record_found.emit(record)
                    found += 1
            except Exception as e:
                self.scan_error.emit(filepath, str(e))

        self.finished.emit(found)

    def _read_record(self, filepath: str) -> Optional[SeriesRecord]:
        uri = f"file:{filepath}?mode=ro"
        con = sqlite3.connect(uri, uri=True)
        con.row_factory = sqlite3.Row
        try:
            cur = con.cursor()

            cur.execute(
                "SELECT id, series, notes_setup, notes_analyst, created, modified "
                "FROM series LIMIT 1"
            )
            row = cur.fetchone()
            if row is None:
                return None
            series_id = row["id"]

            rec = SeriesRecord(
                filepath=filepath,
                filename=os.path.basename(filepath),
                series_id=series_id,
                series_num=str(row["series"] or ""),
                notes_setup=str(row["notes_setup"] or ""),
                notes_analyst=str(row["notes_analyst"] or ""),
                created=_fmt_timestamp(row["created"]),
                modified=_fmt_timestamp(row["modified"]),
            )

            cur.execute(
                "SELECT flyin_name, flyin_location, flyin_date, flyin_analyst "
                "FROM flyin WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            if row:
                rec.flyin_name = str(row["flyin_name"] or "")
                rec.flyin_location = str(row["flyin_location"] or "")
                rec.flyin_date = str(row["flyin_date"] or "")
                rec.flyin_analyst = str(row["flyin_analyst"] or "")

            cur.execute(
                "SELECT pilot, business, street, city, state, zip, phone, email "
                "FROM applicator WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            if row:
                rec.pilot = str(row["pilot"] or "")
                rec.business = str(row["business"] or "")
                rec.street = str(row["street"] or "")
                rec.city = str(row["city"] or "")
                rec.state = str(row["state"] or "")
                rec.zip_code = str(row["zip"] or "")
                rec.phone = str(row["phone"] or "")
                rec.email = str(row["email"] or "")

            cur.execute(
                "SELECT regnum, make, model, wingspan, wingspan_units, winglets "
                "FROM aircraft WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            if row:
                rec.regnum = str(row["regnum"] or "")
                rec.make = str(row["make"] or "")
                rec.model = str(row["model"] or "")
                rec.wingspan = _fmt_with_units(row["wingspan"], row["wingspan_units"])
                rec.winglets = str(row["winglets"] or "")

            # Spray system — keep raw pressure/units for later DSC computation
            cur.execute(
                "SELECT swath, swath_units, rate, rate_units, pressure, pressure_units, "
                "boom_width, boom_width_units, boom_drop, boom_drop_units, "
                "nozzle_spacing, nozzle_spacing_units "
                "FROM spray_system WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            swath_units = cfg.UNIT_FT
            pressure_raw = None
            pressure_units_str = ""
            if row:
                rec.swath = _fmt_with_units(row["swath"], row["swath_units"])
                rec.rate = _fmt_with_units(row["rate"], row["rate_units"])
                rec.pressure = _fmt_with_units(row["pressure"], row["pressure_units"])
                rec.boom_drop = _fmt_with_units(row["boom_drop"], row["boom_drop_units"])
                rec.nozzle_spacing = _fmt_with_units(
                    row["nozzle_spacing"], row["nozzle_spacing_units"]
                )
                swath_units = str(row["swath_units"] or cfg.UNIT_FT)
                pressure_raw = row["pressure"]
                pressure_units_str = str(row["pressure_units"] or "")

            # Passes — aggregate airspeed and height; keep raw values for DSC
            cur.execute(
                "SELECT ground_speed, ground_speed_units, spray_height, spray_height_units "
                "FROM passes WHERE series_id=?",
                (series_id,),
            )
            pass_rows = cur.fetchall()
            rec.pass_count = str(len(pass_rows))

            speeds, speed_unit = [], ""
            heights, height_unit = [], ""
            for pr in pass_rows:
                try:
                    v = float(pr["ground_speed"])
                    if v > 0:
                        speeds.append(v)
                        if not speed_unit:
                            speed_unit = str(pr["ground_speed_units"] or "")
                except (TypeError, ValueError):
                    pass
                try:
                    v = float(pr["spray_height"])
                    if v > 0:
                        heights.append(v)
                        if not height_unit:
                            height_unit = str(pr["spray_height_units"] or "")
                except (TypeError, ValueError):
                    pass

            avg_speed_val = sum(speeds) / len(speeds) if speeds else None
            avg_height_val = sum(heights) / len(heights) if heights else None
            if avg_speed_val is not None:
                rec.avg_speed = _fmt_with_units(f"{avg_speed_val:.0f}", speed_unit)
            if avg_height_val is not None:
                rec.avg_height = _fmt_with_units(f"{avg_height_val:.0f}", height_unit)

            # Nozzle Set 1 (first row)
            cur.execute(
                "SELECT type, size, deflection, quantity FROM nozzles "
                "WHERE series_id=? LIMIT 1",
                (series_id,),
            )
            row = cur.fetchone()
            if row:
                rec.nozzle1_type = str(row["type"] or "")
                rec.nozzle1_size = str(row["size"] or "")
                rec.nozzle1_deflection = str(row["deflection"] or "")
                rec.nozzle1_quantity = str(row["quantity"] or "")

            # Adjusted swaths
            cur.execute(
                "SELECT swath_adjusted FROM series_string WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            if row and row["swath_adjusted"]:
                rec.string_swath_adjusted = _fmt_with_units(
                    str(row["swath_adjusted"]), swath_units
                )

            cur.execute(
                "SELECT swath_adjusted FROM series_spray_card WHERE series_id=?",
                (series_id,),
            )
            row = cur.fetchone()
            if row and row["swath_adjusted"]:
                rec.card_swath_adjusted = _fmt_with_units(
                    str(row["swath_adjusted"]), swath_units
                )

            # Modeled DSC — all nozzle rows + pressure + avg airspeed
            cur.execute(
                "SELECT type, size, deflection, quantity FROM nozzles WHERE series_id=?",
                (series_id,),
            )
            all_nozzles = cur.fetchall()
            airspeed_mph = _to_mph(avg_speed_val, speed_unit)
            pressure_psi = _to_psi(pressure_raw, pressure_units_str)

            if all_nozzles and airspeed_mph and pressure_psi:
                try:
                    model = AtomizationModelMulti()
                    for n in all_nozzles:
                        nozzle_type = str(n["type"] or "")
                        if not nozzle_type:
                            continue
                        model.addNozzleSet(
                            name=nozzle_type,
                            orifice=float(n["size"]) if n["size"] else 0.0,
                            airspeed=airspeed_mph,
                            pressure=pressure_psi,
                            angle=int(float(n["deflection"])) if n["deflection"] else 0,
                            quantity=int(float(n["quantity"])) if n["quantity"] else 1,
                        )
                    rec.modeled_dsc = model.dsc() or ""
                except Exception:
                    rec.modeled_dsc = ""

            # RT CV / B&F CV
            rec.rt_cv, rec.bf_cv = self._compute_cv(cur, series_id, swath_units)

            return rec
        finally:
            con.close()

    def _compute_cv(
        self, cur, series_id: str, swath_units: str
    ) -> tuple[str, str]:
        """Compute RT CV and B&F CV from string pass data. Returns ('', '') on any failure."""
        try:
            cur.execute(
                "SELECT average_center, average_center_method, average_smooth, "
                "average_smooth_window, average_smooth_order, "
                "equalize_integrals, swath_adjusted "
                "FROM series_string WHERE series_id=?",
                (series_id,),
            )
            ss_row = cur.fetchone()
            if not ss_row or not ss_row["swath_adjusted"]:
                return "", ""
            swath_adj = float(ss_row["swath_adjusted"])
            if swath_adj <= 0:
                return "", ""

            cur.execute(
                "SELECT ps.emission_data, ps.include_in_composite, ps.center, "
                "ps.center_method, ps.smooth, ps.smooth_window, ps.smooth_order, "
                "ps.data_loc_units, ps.trim_left, ps.trim_right, ps.trim_vertical, "
                "ps.rebase, p.pass_name, p.pass_number "
                "FROM pass_string ps "
                "JOIN passes p ON p.id = ps.pass_id "
                "WHERE p.series_id = ? "
                "ORDER BY p.pass_number",
                (series_id,),
            )
            pass_rows = cur.fetchall()
            if not pass_rows:
                return "", ""

            passes = []
            for pr in pass_rows:
                if not pr["emission_data"] or not pr["include_in_composite"]:
                    continue
                try:
                    pass_name = str(pr["pass_name"] or f"Pass {pr['pass_number']}")
                    p = Pass(name=pass_name)
                    p.string.data = pd.read_json(StringIO(pr["emission_data"]))
                    p.string.data_loc_units = str(pr["data_loc_units"] or cfg.UNIT_FT)
                    p.string.trim_l = int(pr["trim_left"] or 0)
                    p.string.trim_r = int(pr["trim_right"] or 0)
                    p.string.trim_v = float(pr["trim_vertical"] or 0)
                    p.string.rebase = bool(pr["rebase"])
                    p.string.center = bool(pr["center"])
                    p.string.center_method = str(pr["center_method"] or "")
                    p.string.smooth = bool(pr["smooth"])
                    p.string.smooth_window = float(pr["smooth_window"] or 5)
                    p.string.smooth_order = int(pr["smooth_order"] or 3)
                    p.string.include_in_composite = True
                    passes.append(p)
                except Exception:
                    continue

            if not passes:
                return "", ""

            sds = SeriesDataString(
                passes=passes,
                swath=swath_adj,
                swath_adjusted=swath_adj,
                swath_units=swath_units,
            )
            sds.center = bool(ss_row["average_center"])
            sds.center_method = str(ss_row["average_center_method"] or "")
            sds.smooth = bool(ss_row["average_smooth"])
            sds.smooth_window = float(ss_row["average_smooth_window"] or 5)
            sds.smooth_order = int(ss_row["average_smooth_order"] or 3)
            sds.equalize_integrals = bool(ss_row["equalize_integrals"])

            sds.modifyPatterns()
            average_df = sds.get_average_mod()
            y_label = sds.get_average_y_label()

            if average_df.empty or y_label not in average_df.columns:
                return "", ""

            rt = sds._calcCV(average_df, y_label, swath_adj, mirrorAdjacent=False)
            bf = sds._calcCV(average_df, y_label, swath_adj, mirrorAdjacent=True)
            return f"{rt}%", f"{bf}%"

        except Exception:
            return "", ""


# ---------------------------------------------------------------------------
# Multi-column filter proxy
# ---------------------------------------------------------------------------

class MultiColumnFilterProxy(QSortFilterProxyModel):
    filters_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._col_filters: dict[str, ColumnFilter] = {}

    def set_column_filter(self, key: str, f: Optional[ColumnFilter]):
        if f is None or not f.is_active():
            self._col_filters.pop(key, None)
        else:
            self._col_filters[key] = f
        self.invalidateFilter()
        self.filters_changed.emit()

    def get_column_filter(self, key: str) -> Optional[ColumnFilter]:
        return self._col_filters.get(key)

    def has_column_filter(self, key: str) -> bool:
        return key in self._col_filters

    def has_any_filter(self) -> bool:
        return bool(self._col_filters)

    def clear_all_filters(self):
        if self._col_filters:
            self._col_filters.clear()
            self.invalidateFilter()
            self.filters_changed.emit()

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        if not super().filterAcceptsRow(source_row, source_parent):
            return False
        if not self._col_filters:
            return True
        model = self.sourceModel()
        rec: Optional[SeriesRecord] = model.record_at(source_row)
        if rec is None:
            return False
        for key, col_filter in self._col_filters.items():
            value = str(getattr(rec, key, "") or "")
            if not col_filter.matches(value):
                return False
        return True

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            label = super().headerData(section, orientation, role)
            model = self.sourceModel()
            if model is not None:
                col_def = model.column_def_at(section)
                if col_def is not None and self.has_column_filter(col_def.key):
                    return f"{label} ▼"
        return super().headerData(section, orientation, role)

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:
        model = self.sourceModel()
        col_def = model.column_def_at(left.column()) if model else None
        if col_def is None:
            return super().lessThan(left, right)

        l_str = str(model.data(left, Qt.ItemDataRole.DisplayRole) or "")
        r_str = str(model.data(right, Qt.ItemDataRole.DisplayRole) or "")

        if col_def.filter_type == FilterType.NUMERIC:
            l_num = _parse_leading_float(l_str)
            r_num = _parse_leading_float(r_str)
            if l_num is None and r_num is None:
                return False
            if l_num is None:
                return False  # empty sorts last
            if r_num is None:
                return True
            return l_num < r_num

        if col_def.filter_type == FilterType.DATE:
            l_date = _parse_date(l_str)
            r_date = _parse_date(r_str)
            if l_date is None and r_date is None:
                return False
            if l_date is None:
                return False
            if r_date is None:
                return True
            return l_date < r_date

        return l_str.lower() < r_str.lower()


# ---------------------------------------------------------------------------
# Table model
# ---------------------------------------------------------------------------

class SeriesBrowserTableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._records: list[SeriesRecord] = []
        self._visible: list[ColumnDef] = [
            c for c in _COLUMN_DEFS if c.default_visible
        ]

    def rowCount(self, parent=QModelIndex()) -> int:
        return len(self._records)

    def columnCount(self, parent=QModelIndex()) -> int:
        return len(self._visible)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        rec = self._records[index.row()]
        col = self._visible[index.column()]
        if role == Qt.ItemDataRole.DisplayRole:
            return getattr(rec, col.key, "")
        if role == Qt.ItemDataRole.TextAlignmentRole:
            return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        if role == Qt.ItemDataRole.BackgroundRole and col.key == "modeled_dsc":
            return _DSC_BRUSHES.get(getattr(rec, "modeled_dsc", ""))
        return None

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal:
            if role == Qt.ItemDataRole.DisplayRole:
                return self._visible[section].label
            if role == Qt.ItemDataRole.TextAlignmentRole:
                return int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        return None

    def flags(self, index: QModelIndex) -> Qt.ItemFlag:
        return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable

    def append_record(self, record: SeriesRecord):
        row = len(self._records)
        self.beginInsertRows(QModelIndex(), row, row)
        self._records.append(record)
        self.endInsertRows()

    def clear(self):
        self.beginResetModel()
        self._records.clear()
        self.endResetModel()

    def set_column_visible(self, key: str, visible: bool):
        self.beginResetModel()
        if visible:
            current_keys = {c.key for c in self._visible}
            current_keys.add(key)
            self._visible = [c for c in _COLUMN_DEFS if c.key in current_keys]
        else:
            self._visible = [c for c in self._visible if c.key != key]
        self.endResetModel()

    def record_at(self, row: int) -> Optional[SeriesRecord]:
        if 0 <= row < len(self._records):
            return self._records[row]
        return None

    def column_def_at(self, col: int) -> Optional[ColumnDef]:
        if 0 <= col < len(self._visible):
            return self._visible[col]
        return None

    def visible_columns(self) -> list[ColumnDef]:
        return list(self._visible)


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class DatabaseBrowserWindow(QDialog):
    open_requested = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setWindowTitle("Database Browser")
        self.resize(1280, 720)
        self.setMinimumSize(800, 480)
        self.setWindowFlags(
            Qt.WindowType.Window
            | Qt.WindowType.WindowCloseButtonHint
            | Qt.WindowType.WindowMinMaxButtonsHint
        )

        self._directory = ""
        self._scanner_thread: Optional[QThread] = None
        self._scanner: Optional[DatabaseScanner] = None
        self._scan_complete = False

        self._model = SeriesBrowserTableModel()
        self._proxy = MultiColumnFilterProxy()
        self._proxy.setSourceModel(self._model)
        self._proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._proxy.setFilterKeyColumn(-1)
        self._proxy.filters_changed.connect(self._update_status)

        self._setup_ui()
        self.show()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 8)
        root.setSpacing(6)

        # Directory picker row
        top = QHBoxLayout()
        top.setSpacing(8)
        self._btn_choose = QPushButton("Choose Directory…")
        self._btn_choose.clicked.connect(self._choose_directory)
        self._lbl_dir = QLabel("No directory selected")
        self._lbl_dir.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._lbl_dir.setStyleSheet("color: gray; font-style: italic;")
        self._lbl_dir.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self._btn_rescan = QPushButton("Rescan")
        self._btn_rescan.setEnabled(False)
        self._btn_rescan.clicked.connect(self._start_scan)
        top.addWidget(self._btn_choose)
        top.addWidget(self._lbl_dir, 1)
        top.addWidget(self._btn_rescan)
        root.addLayout(top)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFrameShadow(QFrame.Shadow.Sunken)
        root.addWidget(sep)

        # Filter + action bar
        bar = QHBoxLayout()
        bar.setSpacing(8)

        self._filter_input = QLineEdit()
        self._filter_input.setPlaceholderText("Right-Click Column Headers or Type Here to Filter Series…")
        self._filter_input.setClearButtonEnabled(True)
        self._filter_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._filter_input.textChanged.connect(self._proxy.setFilterFixedString)
        self._filter_input.textChanged.connect(lambda _: self._update_status())

        self._btn_columns = QToolButton()
        self._btn_columns.setText("Columns  ▾")
        self._btn_columns.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        col_menu = QMenu(self._btn_columns)
        self._col_actions: dict[str, QAction] = {}
        for item in ALL_COLUMNS:
            if isinstance(item, ColumnSection):
                col_menu.addSection(item.title)
            else:
                action = QAction(item.label, col_menu)
                action.setCheckable(True)
                action.setChecked(item.default_visible)
                action.setData(item.key)
                key = item.key
                action.toggled.connect(
                    lambda checked, k=key: self._on_column_toggled(k, checked)
                )
                col_menu.addAction(action)
                self._col_actions[item.key] = action
        self._btn_columns.setMenu(col_menu)

        self._btn_export = QPushButton("Export to Excel…")
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._export_excel)

        self._btn_reveal = QPushButton("Reveal in Finder…")
        self._btn_reveal.setEnabled(False)
        self._btn_reveal.clicked.connect(self._reveal_selected)

        self._btn_open = QPushButton("Open in AccuPatt")
        self._btn_open.setEnabled(False)
        self._btn_open.setDefault(False)
        self._btn_open.clicked.connect(self._open_selected)

        bar.addWidget(self._filter_input, 1)
        bar.addWidget(self._btn_columns)
        bar.addSpacing(4)
        bar.addWidget(self._btn_export)
        bar.addWidget(self._btn_reveal)
        bar.addWidget(self._btn_open)
        root.addLayout(bar)

        # Table
        self._table = QTableView()
        self._table.setModel(self._proxy)
        self._table.setSortingEnabled(True)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setShowGrid(False)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setHighlightSections(False)
        hdr.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        hdr.customContextMenuRequested.connect(self._on_header_context_menu)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(22)
        self._table.doubleClicked.connect(lambda _: self._on_double_click())
        self._table.selectionModel().selectionChanged.connect(self._on_selection_changed)
        root.addWidget(self._table, 1)

        # Status row
        status_row = QHBoxLayout()
        status_row.setSpacing(8)
        self._status_label = QLabel("Choose a directory to begin.")
        self._status_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._path_label = QLabel("")
        self._path_label.setStyleSheet("color: gray; font-size: 11px;")
        self._path_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self._path_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        status_row.addWidget(self._status_label, 1)
        status_row.addWidget(self._path_label)
        root.addLayout(status_row)

    # --- Directory and scan ---

    @pyqtSlot()
    def _choose_directory(self):
        directory = QFileDialog.getExistingDirectory(
            parent=self,
            caption="Select Directory to Search",
            directory=self._directory or os.path.expanduser("~"),
            options=QFileDialog.Option.ShowDirsOnly,
        )
        if directory:
            self._directory = directory
            self._lbl_dir.setText(directory)
            self._lbl_dir.setStyleSheet("")
            self._btn_rescan.setEnabled(True)
            self._start_scan()

    @pyqtSlot()
    def _start_scan(self):
        if not self._directory:
            return

        if self._scanner is not None:
            self._scanner.cancel()
        if self._scanner_thread is not None and self._scanner_thread.isRunning():
            self._scanner_thread.quit()
            self._scanner_thread.wait(1000)

        self._scan_complete = False
        self._model.clear()
        self._status_label.setText("Scanning…")
        self._btn_export.setEnabled(False)
        self._btn_open.setEnabled(False)

        self._scanner = DatabaseScanner(self._directory)
        self._scanner_thread = QThread()
        self._scanner.moveToThread(self._scanner_thread)

        self._scanner_thread.started.connect(self._scanner.run)
        self._scanner.progress.connect(self._on_scan_progress)
        self._scanner.record_found.connect(self._model.append_record)
        self._scanner.finished.connect(self._on_scan_finished)
        self._scanner.finished.connect(self._scanner_thread.quit)

        self._scanner_thread.start()

    @pyqtSlot(int, int)
    def _on_scan_progress(self, current: int, total: int):
        if total > 0:
            self._status_label.setText(f"Scanning… {current} of {total}")

    @pyqtSlot(int)
    def _on_scan_finished(self, count: int):
        self._scan_complete = True
        self._btn_export.setEnabled(count > 0)
        self._table.resizeColumnsToContents()
        self._table.horizontalHeader().setStretchLastSection(True)
        self._update_status()

    # --- Column visibility ---

    def _on_column_toggled(self, key: str, checked: bool):
        self._model.set_column_visible(key, checked)
        if checked:
            self._table.resizeColumnsToContents()
            self._table.horizontalHeader().setStretchLastSection(True)

    # --- Header right-click: column filters ---

    @pyqtSlot(QPoint)
    def _on_header_context_menu(self, pos: QPoint):
        header = self._table.horizontalHeader()
        logical_col = header.logicalIndexAt(pos)
        if logical_col < 0:
            return
        col_def = self._model.column_def_at(logical_col)
        if col_def is None:
            return

        menu = QMenu(self)
        has_filter = self._proxy.has_column_filter(col_def.key)
        act_filter = menu.addAction(f"Filter '{col_def.label}'…")
        act_clear_col = menu.addAction("Clear This Filter")
        act_clear_col.setEnabled(has_filter)
        menu.addSeparator()
        act_clear_all = menu.addAction("Clear All Column Filters")
        act_clear_all.setEnabled(self._proxy.has_any_filter())

        action = menu.exec(header.mapToGlobal(pos))
        if action == act_filter:
            self._open_filter_dialog(col_def)
        elif action == act_clear_col:
            self._proxy.set_column_filter(col_def.key, None)
        elif action == act_clear_all:
            self._proxy.clear_all_filters()

    def _open_filter_dialog(self, col_def: ColumnDef):
        existing = self._proxy.get_column_filter(col_def.key)
        if col_def.filter_type == FilterType.TEXT:
            dlg = TextFilterDialog(col_def.label, existing, parent=self)
        elif col_def.filter_type == FilterType.NUMERIC:
            dlg = NumericRangeDialog(col_def.label, existing, parent=self)
        else:
            dlg = DateRangeDialog(col_def.label, existing, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._proxy.set_column_filter(col_def.key, dlg.result_filter())

    # --- Selection ---

    @pyqtSlot()
    def _on_selection_changed(self):
        rows = self._table.selectionModel().selectedRows()
        single = len(rows) == 1
        self._btn_reveal.setEnabled(single)
        self._btn_open.setEnabled(single)
        if single:
            source_row = self._proxy.mapToSource(rows[0]).row()
            rec = self._model.record_at(source_row)
            if rec and self._directory:
                try:
                    rel = os.path.relpath(rec.filepath, self._directory)
                except ValueError:
                    rel = rec.filepath
                self._path_label.setText(rel)
            else:
                self._path_label.setText("")
        else:
            self._path_label.setText("")

    @pyqtSlot()
    def _reveal_selected(self):
        rows = self._table.selectionModel().selectedRows()
        if len(rows) != 1:
            return
        source_row = self._proxy.mapToSource(rows[0]).row()
        rec = self._model.record_at(source_row)
        if rec and os.path.exists(rec.filepath):
            subprocess.run(["open", "-R", rec.filepath])

    @pyqtSlot()
    def _on_double_click(self):
        self._open_selected()

    @pyqtSlot()
    def _open_selected(self):
        rows = self._table.selectionModel().selectedRows()
        if len(rows) != 1:
            return
        source_row = self._proxy.mapToSource(rows[0]).row()
        rec = self._model.record_at(source_row)
        if rec is None:
            return
        if not os.path.exists(rec.filepath):
            QMessageBox.warning(self, "File Not Found", f"Could not locate:\n{rec.filepath}")
            return
        self.open_requested.emit(rec.filepath)

    # --- Export ---

    @pyqtSlot()
    def _export_excel(self):
        proxy_rows = list(range(self._proxy.rowCount()))
        if not proxy_rows:
            return

        records: list[SeriesRecord] = []
        for pr in proxy_rows:
            source_idx = self._proxy.mapToSource(self._proxy.index(pr, 0))
            rec = self._model.record_at(source_idx.row())
            if rec:
                records.append(rec)
        if not records:
            return

        save_path, _ = QFileDialog.getSaveFileName(
            parent=self,
            caption="Export to Excel",
            directory=self._directory or os.path.expanduser("~"),
            filter="Excel Workbook (*.xlsx)",
        )
        if not save_path:
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        visible_cols = self._model.visible_columns()
        wb = Workbook()
        ws = wb.active
        ws.title = "Series"

        for col_idx, col_def in enumerate(visible_cols, 1):
            cell = ws.cell(1, col_idx, col_def.label)
            cell.style = "Pandas"

        for row_idx, rec in enumerate(records, 2):
            for col_idx, col_def in enumerate(visible_cols, 1):
                ws.cell(row_idx, col_idx, getattr(rec, col_def.key, ""))

        try:
            wb.save(save_path)
            n = len(records)
            self._status_label.setText(
                f"Exported {n} {'row' if n == 1 else 'rows'} to {os.path.basename(save_path)}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save file:\n{e}")

    # --- Status ---

    def _update_status(self):
        if not self._scan_complete:
            return
        total = self._model.rowCount()
        shown = self._proxy.rowCount()
        if total == 0:
            self._status_label.setText("No AccuPatt series found in directory.")
            return
        has_global = bool(self._filter_input.text().strip())
        has_col = self._proxy.has_any_filter()
        if shown == total:
            self._status_label.setText(f"{total} series found")
        elif has_global and has_col:
            self._status_label.setText(
                f"{shown} of {total} series shown  (search + column filters active)"
            )
        elif has_col:
            n = len(self._proxy._col_filters)
            self._status_label.setText(
                f"{shown} of {total} series shown  "
                f"({n} column {'filter' if n == 1 else 'filters'} active)"
            )
        else:
            self._status_label.setText(
                f"{shown} of {total} series shown  (search active)"
            )

    # --- Lifecycle ---

    def closeEvent(self, event):
        if self._scanner is not None:
            self._scanner.cancel()
        if self._scanner_thread is not None and self._scanner_thread.isRunning():
            self._scanner_thread.quit()
            self._scanner_thread.wait(2000)
        super().closeEvent(event)
