#!/usr/bin/env python3
"""
One-time generator. Reads the USDA AATRU atomization-model spreadsheet and
emits one Python file per nozzle into
accupatt/helpers/atomizationModel/nozzles/.

Each emitted file defines a single `NOZZLE: Nozzle` constant. Re-running the
script overwrites those files in place; nothing else is touched.

    python tools/extract_atomization_model.py [path/to/2026-USDA.xlsx]
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]
NOZZLES_DIR = REPO_ROOT / "accupatt" / "helpers" / "atomizationModel" / "nozzles"

EXCEL_TO_MODEL_NAME: dict[str, str] = {
    "CP11TT 20° Flat Fan": "CP11TT 20°FF",
    "CP11TT 40° Flat Fan": "CP11TT 40°FF",
    "CP11TT 60° Flat Fan": "CP11TT 60°FF",
    "CP11TT 80° Flat Fan": "CP11TT 80°FF",
    "CP11TT 110° Flat Fan": "CP11TT 110°FF",
    "CP11TT Straight Stream": "CP11TT SS",
    "CP-03": "CP03",
    "CP-09 Deflection Only": "CP09 Deflection",
    "CP-09 Straight Stream": "CP09 SS",
    "CP-01-03": "CP-01-03",
    "CP-07-3E Deflection Only": "CP-07-3E Deflection",
    "CP-07-3E Straight Stream": "CP-07-3E SS",
    "Steel Disc Core 45": "Steel Disc Core 45",
    "Steel Disc Core Straight Stream": "Steel Disc Core SS",
    "Ceramic Disc Core 45": "Ceramic Disc Core 45",
    "Ceramic Disc Core Straight Stream": "Ceramic Disc Core SS",
    "Standard 40° Flat Fan": "Standard 40°FF",
    "Standard 80° Flat Fan": "Standard 80°FF",
    "TriSet Deflection Only": "Davidon TriSet Deflection",
    "TriSet Straight Stream": "Davidon TriSet SS",
    "TeeJet SS": "TeeJet SS",
    "TeeJet H1 4U": "TeeJet H1 4U",
    "CAS LF-5 Straight Stream": "CAS LF-5 SS",
    "CAS LF-5 Deflection Only": "CAS LF-5 Deflection",
    "AFS Straight Stream": "AFS SS",
    "AccuFlow Single Row": "AccuFlow Single Row",
    "AccuFlow Double Row": "AccuFlow Double Row",
}

ACCUFLOW_EXCEL_NAMES = {"AccuFlow Single Row", "AccuFlow Double Row"}

# Row anchors in the "Model Parameters with Tables" sheet.
# Each "Params" anchor is the title row; the column-header row is anchor+1;
# data starts at anchor+2 and runs until the next blank A-cell.
PARAMS_ANCHORS = {
    "HS": {
        "dv10": 93,
        "dv50": 149,
        "dv90": 206,
        "v100": 263,
        "v200": 321,
    },
    "LS": {
        "dv10": 121,
        "dv50": 177,
        "dv90": 234,
        "v100": 291,
        "v200": 349,
    },
}
# CCD Values: shared title row at 65, data starts at 67.
# HS uses cols A-I, LS uses cols L-T.
CCD_ANCHOR = 65
ORIFICE_ANCHOR = 379          # Orifice header at 379, valid values at 381+
ANGLE_ANCHORS = {"HS": 417, "LS": 446}
AIRSPEED_ANCHORS = {"HS": 476, "LS": 505}
PRESSURE_ANCHORS = {"HS": 535, "LS": 565}

COEFF_FIELDS = (
    "intercept",
    "orifice",
    "airspeed",
    "pressure",
    "angle",
    "orifice_airspeed",
    "orifice_pressure",
    "airspeed_pressure",
    "orifice_angle",
    "airspeed_angle",
    "pressure_angle",
    "orifice_squared",
    "airspeed_squared",
    "pressure_squared",
    "angle_squared",
)


def slug(name: str) -> str:
    """Convert a nozzle name into a safe Python module filename."""
    s = unicodedata.normalize("NFKD", name)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.replace("°", "deg")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").lower()
    return s


def fmt_num(x: float | int) -> str:
    """Format a numeric value compactly but losslessly enough for review."""
    if isinstance(x, int):
        return str(x)
    if x == int(x):
        return f"{int(x)}"
    return f"{x!r}"


def read_table(
    ws: Worksheet, start_row: int, name_col: int, value_cols: range
) -> dict[str, tuple]:
    """Read a name -> tuple-of-values table starting at start_row. Stops at first blank name."""
    out: dict[str, tuple] = {}
    r = start_row
    while True:
        name = ws.cell(row=r, column=name_col).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        values = tuple(ws.cell(row=r, column=c).value for c in value_cols)
        out[name] = values
        r += 1
    return out


def read_value_list(ws: Worksheet, start_row: int, value_col_start: int) -> dict[str, tuple]:
    """
    Read a 'valid values' style table where each row is:
        col A = nozzle name, col value_col_start..N = values (None where unused)
    """
    out: dict[str, tuple] = {}
    r = start_row
    while True:
        name = ws.cell(row=r, column=1).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        vals = []
        c = value_col_start
        while True:
            v = ws.cell(row=r, column=c).value
            # values stretch only over a fixed window; stop at first None
            if v is None:
                break
            vals.append(v)
            c += 1
        out[name] = tuple(vals)
        r += 1
    return out


def first_row_after(ws: Worksheet, anchor_row: int, label: str) -> int:
    """Return the row index of the first nozzle data row after a title anchor."""
    # Title row is anchor_row; column header row is anchor_row + 1; data starts at anchor_row + 2.
    return anchor_row + 2


def parse_coefficients_table(ws: Worksheet, anchor: int) -> dict[str, dict]:
    """Read 15 polynomial coefficients per nozzle from a Params table."""
    start = first_row_after(ws, anchor, "params")
    raw = read_table(ws, start, name_col=1, value_cols=range(2, 17))
    return {name: dict(zip(COEFF_FIELDS, vals)) for name, vals in raw.items()}


def parse_ccd_table(ws: Worksheet, col_offset: int) -> dict[str, dict]:
    """Read CCD values (Orf Sub, Orf Div, AS Sub, AS Div, Press Sub, Press Div, Ang Sub, Ang Div)."""
    # Header row is 66 (CCD_ANCHOR + 1), data starts at 67.
    start = CCD_ANCHOR + 2
    name_col = 1 + col_offset
    val_cols = range(2 + col_offset, 10 + col_offset)
    out: dict[str, dict] = {}
    r = start
    while True:
        name = ws.cell(row=r, column=name_col).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        out[name] = {
            "orifice_sub": ws.cell(row=r, column=val_cols.start + 0).value,
            "orifice_div": ws.cell(row=r, column=val_cols.start + 1).value,
            "airspeed_sub": ws.cell(row=r, column=val_cols.start + 2).value,
            "airspeed_div": ws.cell(row=r, column=val_cols.start + 3).value,
            "pressure_sub": ws.cell(row=r, column=val_cols.start + 4).value,
            "pressure_div": ws.cell(row=r, column=val_cols.start + 5).value,
            "angle_sub": ws.cell(row=r, column=val_cols.start + 6).value,
            "angle_div": ws.cell(row=r, column=val_cols.start + 7).value,
        }
        r += 1
    return out


def parse_orifice_values(ws: Worksheet) -> dict[str, tuple]:
    """Orifice values per nozzle. Row layout: A=name, B=range-text, C..=values."""
    out: dict[str, tuple] = {}
    r = ORIFICE_ANCHOR + 2
    while True:
        name = ws.cell(row=r, column=1).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        vals: list[float] = []
        c = 3
        while True:
            v = ws.cell(row=r, column=c).value
            if v is None:
                break
            vals.append(v)
            c += 1
        out.setdefault(name, tuple(vals))  # first occurrence wins; some names repeat
        r += 1
    return out


def parse_angle_table(ws: Worksheet, anchor: int) -> dict[str, tuple[str, tuple]]:
    """Angle values + descriptor per nozzle. Row: A=name, B=descriptor text, C..=values."""
    out: dict[str, tuple[str, tuple]] = {}
    r = anchor + 2
    while True:
        name = ws.cell(row=r, column=1).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        descriptor = ws.cell(row=r, column=2).value or "Nozzle Angle"
        vals: list[float] = []
        c = 3
        while True:
            v = ws.cell(row=r, column=c).value
            if v is None or (isinstance(v, str) and "to" in v.lower()):
                break
            vals.append(v)
            c += 1
        out[name] = (descriptor, tuple(vals))
        r += 1
    return out


def parse_airspeed_table(ws: Worksheet, anchor: int) -> dict[str, tuple]:
    """Airspeed values per nozzle. Row: A=name, B..=values, last col = range text."""
    out: dict[str, tuple] = {}
    r = anchor + 2
    while True:
        name = ws.cell(row=r, column=1).value
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        vals: list[float] = []
        c = 2
        while True:
            v = ws.cell(row=r, column=c).value
            if v is None or isinstance(v, str):
                break
            vals.append(v)
            c += 1
        out[name] = tuple(vals)
        r += 1
    return out


def parse_pressure_table(ws: Worksheet, anchor: int) -> dict[str, tuple]:
    """Pressure values per nozzle. Row: A=name, B..=values, last col = range text."""
    return parse_airspeed_table(ws, anchor)  # identical layout


def parse_flow_rates(ws: Worksheet) -> tuple[dict, dict]:
    """
    Returns (flat_flows, accuflow_flows).

    flat_flows[excel_name][orifice_size] = (a, b)
    accuflow_flows[excel_name][orifice_size][restrictor_size] = (a, b)
    """
    flat: dict[str, dict[float, tuple[float, float]]] = defaultdict(dict)
    accu: dict[str, dict[float, dict[float, tuple[float, float]]]] = defaultdict(
        lambda: defaultdict(dict)
    )

    r = 1
    max_row = ws.max_row
    while r <= max_row:
        name = ws.cell(row=r, column=1).value
        if isinstance(name, str) and name in EXCEL_TO_MODEL_NAME:
            if name in ACCUFLOW_EXCEL_NAMES:
                orifice = ws.cell(row=r, column=2).value
                restrictor = ws.cell(row=r, column=3).value
                a = ws.cell(row=r, column=4).value
                b = ws.cell(row=r, column=5).value
                if (
                    orifice is not None
                    and restrictor is not None
                    and a is not None
                    and b is not None
                ):
                    accu[name][orifice][restrictor] = (a, b)
            else:
                orifice = ws.cell(row=r, column=2).value
                a = ws.cell(row=r, column=3).value
                b = ws.cell(row=r, column=4).value
                if orifice is not None and a is not None and b is not None:
                    flat[name].setdefault(orifice, (a, b))
        r += 1

    return dict(flat), dict(accu)


def emit_coefficient_block(name: str, c: dict, indent: str) -> str:
    inner = ",\n".join(
        f"{indent}    {field}={fmt_num(c[field])}" for field in COEFF_FIELDS
    )
    return f"Coefficients(\n{inner},\n{indent})"


def emit_ccd(c: dict, indent: str) -> str:
    fields = (
        "orifice_sub",
        "orifice_div",
        "airspeed_sub",
        "airspeed_div",
        "pressure_sub",
        "pressure_div",
        "angle_sub",
        "angle_div",
    )
    inner = ", ".join(f"{f}={fmt_num(c[f])}" for f in fields)
    return f"CCD({inner})"


def emit_tuple(values: tuple) -> str:
    if not values:
        return "()"
    if len(values) == 1:
        return f"({fmt_num(values[0])},)"
    return "(" + ", ".join(fmt_num(v) for v in values) + ")"


def emit_orifice(size: float, flow_data, indent: str) -> str:
    if isinstance(flow_data, dict):
        items = ",\n".join(
            f"{indent}        {fmt_num(r)}: FlowRate(a={fmt_num(a)}, b={fmt_num(b)})"
            for r, (a, b) in sorted(flow_data.items())
        )
        flow_repr = "{\n" + items + f",\n{indent}    }}"
    else:
        a, b = flow_data
        flow_repr = f"FlowRate(a={fmt_num(a)}, b={fmt_num(b)})"
    return f"{indent}Orifice(size={fmt_num(size)}, flow={flow_repr})"


def render_nozzle(
    model_name: str,
    angle_description: str,
    orifice_records: list[tuple[float, object]],
    models: list[dict],
) -> str:
    """Render a single nozzle's NOZZLE = Nozzle(...) Python source."""
    orifice_block = ",\n".join(
        emit_orifice(size, flow, "    ") for size, flow in orifice_records
    )

    model_blocks = []
    for m in models:
        ccd_s = emit_ccd(m["ccd"], "        ")
        coeffs = "\n".join(
            f"            {key}={emit_coefficient_block(key, m[key], '            ')},"
            for key in ("coeff_dv10", "coeff_dv50", "coeff_dv90", "coeff_v100", "coeff_v200")
        )
        model_blocks.append(
            f"""        Model(
            name={m["name"]!r},
            airspeeds={emit_tuple(m["airspeeds"])},
            angles={emit_tuple(m["angles"])},
            pressures={emit_tuple(m["pressures"])},
            ccd={ccd_s},
{coeffs}
        )"""
        )

    models_block = ",\n".join(model_blocks)

    return f'''"""Generated by tools/extract_atomization_model.py. Do not edit by hand."""

from ..types import CCD, Coefficients, FlowRate, Model, Nozzle, Orifice


NOZZLE = Nozzle(
    name={model_name!r},
    angle_description={angle_description!r},
    orifices=(
{orifice_block},
    ),
    models=(
{models_block},
    ),
)
'''


def coalesce_descriptor(descriptors: list[str | None]) -> str:
    """Pick the most informative angle descriptor seen across LS/HS."""
    for d in descriptors:
        if d and d.lower() not in {"no deflection", ""}:
            return d
    for d in descriptors:
        if d:
            return d
    return "Nozzle Angle"


def build() -> dict[str, str]:
    """Extract everything and return {filename: source_code}."""
    args = argparse.ArgumentParser()
    args.add_argument(
        "xlsx",
        nargs="?",
        default="/Users/mattgill/Downloads/2026-USDA-ARS-AATRU-Atomization-Models.xlsx",
    )
    parsed = args.parse_args()

    wb = load_workbook(parsed.xlsx, data_only=True)
    params_ws = wb["Model Parameters with Tables"]
    flows_ws = wb["Nozzle Flow Rates with Tables"]

    # --- model parameters (per regime) ---
    regime_data = {}
    for regime in ("HS", "LS"):
        anchors = PARAMS_ANCHORS[regime]
        coeffs_per_param = {
            "coeff_dv10": parse_coefficients_table(params_ws, anchors["dv10"]),
            "coeff_dv50": parse_coefficients_table(params_ws, anchors["dv50"]),
            "coeff_dv90": parse_coefficients_table(params_ws, anchors["dv90"]),
            "coeff_v100": parse_coefficients_table(params_ws, anchors["v100"]),
            "coeff_v200": parse_coefficients_table(params_ws, anchors["v200"]),
        }
        ccd = parse_ccd_table(params_ws, col_offset=0 if regime == "HS" else 11)
        angles = parse_angle_table(params_ws, ANGLE_ANCHORS[regime])
        airspeeds = parse_airspeed_table(params_ws, AIRSPEED_ANCHORS[regime])
        pressures = parse_pressure_table(params_ws, PRESSURE_ANCHORS[regime])
        regime_data[regime] = {
            "coeffs_per_param": coeffs_per_param,
            "ccd": ccd,
            "angles": angles,
            "airspeeds": airspeeds,
            "pressures": pressures,
        }

    orifices = parse_orifice_values(params_ws)
    flat_flows, accu_flows = parse_flow_rates(flows_ws)

    # --- collect nozzles to emit ---
    all_excel_names: set[str] = set()
    for regime in ("HS", "LS"):
        all_excel_names |= set(regime_data[regime]["coeffs_per_param"]["coeff_dv10"].keys())

    unknown = all_excel_names - set(EXCEL_TO_MODEL_NAME)
    if unknown:
        print(f"WARNING: unmapped Excel nozzle names: {sorted(unknown)}", file=sys.stderr)

    out: dict[str, str] = {}
    for excel_name in sorted(all_excel_names):
        if excel_name not in EXCEL_TO_MODEL_NAME:
            continue
        model_name = EXCEL_TO_MODEL_NAME[excel_name]

        # Build per-regime Model entries
        models = []
        descriptors = []
        for regime, regime_label in (("LS", "low-speed"), ("HS", "high-speed")):
            data = regime_data[regime]
            if excel_name not in data["coeffs_per_param"]["coeff_dv10"]:
                continue
            ccd = data["ccd"].get(excel_name)
            if not ccd:
                print(f"WARNING: {excel_name} missing CCD for {regime}", file=sys.stderr)
                continue
            desc, angle_vals = data["angles"].get(excel_name, ("Nozzle Angle", ()))
            descriptors.append(desc)
            models.append(
                {
                    "name": regime_label,
                    "airspeeds": data["airspeeds"].get(excel_name, ()),
                    "angles": angle_vals,
                    "pressures": data["pressures"].get(excel_name, ()),
                    "ccd": ccd,
                    "coeff_dv10": data["coeffs_per_param"]["coeff_dv10"][excel_name],
                    "coeff_dv50": data["coeffs_per_param"]["coeff_dv50"][excel_name],
                    "coeff_dv90": data["coeffs_per_param"]["coeff_dv90"][excel_name],
                    "coeff_v100": data["coeffs_per_param"]["coeff_v100"][excel_name],
                    "coeff_v200": data["coeffs_per_param"]["coeff_v200"][excel_name],
                }
            )

        if not models:
            continue

        angle_description = coalesce_descriptor(descriptors)

        # Build orifice records. The spreadsheet's Orifice table omits some
        # nozzles (e.g. Ceramic Disc Core Straight Stream); fall back to the
        # orifices we have flow data for.
        orifice_sizes = orifices.get(excel_name)
        if not orifice_sizes:
            if excel_name in ACCUFLOW_EXCEL_NAMES:
                orifice_sizes = tuple(sorted(accu_flows.get(excel_name, {}).keys()))
            else:
                orifice_sizes = tuple(sorted(flat_flows.get(excel_name, {}).keys()))
            if orifice_sizes:
                print(
                    f"NOTE: {excel_name} not in Orifice table; using flow-rate orifices "
                    f"{list(orifice_sizes)}",
                    file=sys.stderr,
                )
        is_accu = excel_name in ACCUFLOW_EXCEL_NAMES
        orifice_records: list[tuple[float, object]] = []
        for size in orifice_sizes:
            if is_accu:
                per_restrictor = accu_flows.get(excel_name, {}).get(size, {})
                if not per_restrictor:
                    print(
                        f"WARNING: {excel_name} missing AccuFlow flow for orifice {size}",
                        file=sys.stderr,
                    )
                    continue
                orifice_records.append((size, dict(per_restrictor)))
            else:
                pair = flat_flows.get(excel_name, {}).get(size)
                if pair is None:
                    # fall back to closest size within 0.005 tolerance
                    candidates = flat_flows.get(excel_name, {})
                    if candidates:
                        nearest = min(candidates.keys(), key=lambda s: abs(s - size))
                        if abs(nearest - size) <= max(0.005, 0.05 * size):
                            pair = candidates[nearest]
                            print(
                                f"NOTE: {excel_name} orifice {size} using flow data from "
                                f"closest match {nearest}",
                                file=sys.stderr,
                            )
                if pair is None:
                    print(
                        f"WARNING: {excel_name} orifice {size} has no flow data; skipping",
                        file=sys.stderr,
                    )
                    continue
                orifice_records.append((size, pair))

        if not orifice_records:
            print(f"WARNING: {excel_name} has no usable orifices; skipping", file=sys.stderr)
            continue

        filename = slug(model_name) + ".py"
        out[filename] = render_nozzle(model_name, angle_description, orifice_records, models)

    return out


def main() -> int:
    files = build()
    NOZZLES_DIR.mkdir(parents=True, exist_ok=True)
    # Remove old generated files (anything except __init__.py)
    for p in NOZZLES_DIR.glob("*.py"):
        if p.name != "__init__.py":
            p.unlink()
    for name, src in sorted(files.items()):
        (NOZZLES_DIR / name).write_text(src)
        print(f"wrote {name}")
    print(f"wrote {len(files)} nozzles")
    return 0


if __name__ == "__main__":
    sys.exit(main())
